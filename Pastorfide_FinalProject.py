import tkinter as tk
from tkinter import ttk,messagebox
import openpyxl as op

#SALON RESERVATION SYSTEM
#Functions

def display():
    workbook = op.load_workbook("Pastorfide_Database.xlsx")
    sheet = workbook.active

    for content in table.get_children():
        table.delete(content)

    for row in sheet.iter_rows(min_row=2,values_only=True):
        table.insert("",tk.END,values=row)    

def select_record(event):
    selected = table.focus()
    values = table.item(selected, "values")

    if not values:
        return
    
    cname_entry.delete(0,tk.END)
    cnumber_entry.delete(0,tk.END)
    service_list.delete(0,tk.END)
    date_list.delete(0,tk.END)
    time_list.delete(0,tk.END)

    cname_entry.insert(0,values[1])
    cnumber_entry.insert(0,values[2])
    service_list.insert(0,values[3])
    date_list.insert(0,values[4])
    time_list.insert(0,values[5])

def validation():
    customer = cname_entry.get()
    contact = cnumber_entry.get()
    service = service_list.get()
    date = date_list.get()
    time = time_list.get()

    if not customer or not contact or not service or not date or not time:
        messagebox.showerror("Error", "All fields are required  to answer")
        return False
    return True
  

def appending():
    if not validation():
        return
    customer = cname_entry.get().title()
    contact = cnumber_entry.get()
    service = service_list.get()
    date = date_list.get()
    time = time_list.get()


    if not contact.isdigit() or len(contact) != 11:
        messagebox.showerror(
        "Error",
        "Contact must be 11 digit number."
    )
        return
    
    if time == "BREAK TIME":
        messagebox.showerror("Sorry","Break time po namin")
        return False
    
    workbook = op.load_workbook("Pastorfide_Database.xlsx")
    sheet=workbook.active

    new_id= sheet.max_row

    sheet.append([new_id,customer,contact,service,date,time])
    workbook.save("Pastorfide_Database.xlsx")

    messagebox.showinfo("Success","Record added successfully!")
    
    display()

def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error","Select a record first.")
        return
    if not validation():
        return
    
    values = table.item(selected,"values")
    record_id = values[0]

    customer = cname_entry.get().title()
    contact = cnumber_entry.get()
    service = service_list.get()
    date = date_list.get()
    time = time_list.get()

    if time == "BREAK TIME":
        messagebox.showerror("Sorry","Break time po namin")
        return False


    workbook = op.load_workbook("Pastorfide_Database.xlsx")
    sheet = workbook.active

    for rows in sheet.iter_rows(min_row=2):
        if str(rows[0].value) == str(record_id):
            rows[1].value = customer
            rows[2].value = contact
            rows [3].value = service
            rows [4].value = date
            rows [5].value = time

    
    workbook.save("Pastorfide_Database.xlsx")
    messagebox.showinfo("Success","Record updated successfully!")
    display()

def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error","Select a record first!")
        return
    values = table.item(selected,"values")
    record_id = values[0]

    confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete this record?")
    
    if not confirm:
        return

    workbook = op.load_workbook("Pastorfide_Database.xlsx")
    sheet = workbook.active

    for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
        if str(row[0].value) == str(record_id):
            sheet.delete_rows(i)
            break

    workbook.save("Pastorfide_Database.xlsx")
    messagebox.showinfo("Success","Record deleted successfully")
    display()

window = tk.Tk()
window.title("Salon Reservation System")
window.configure(bg="lightblue")

# Form Title
title = tk.Label(window, text="Salon Reservation System", font=("Times New Roman", 14, "bold"), bg="lightblue")
title.grid(row=0, column=0, columnspan=6)

# Frame
genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=20, pady=20)

# Customer Name Entry
cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=0, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = tk.Label(genframe, text="Customer Name", font=("Poppins", 10, "italic"), bg="lightblue")
cname_label.grid(row=2, column=1, columnspan=2)

# Contact Number Entry
cnumber_entry = tk.Entry(genframe, font=("Poppins", 12))
cnumber_entry.grid(row=0, column=3, columnspan=2, padx=10, pady=(10, 0))

contact_label = tk.Label(genframe, text="Contact", font=("Poppins", 10, "italic"), bg="lightblue")
contact_label.grid(row=2, column=3, columnspan=2)

# Service Entry
service_list = ttk.Combobox(genframe,values=["",
                                            "Hair Coloring",
                                            "Hair Rebonding",
                                            "Keratin Treatment",
                                            "Haircut",
                                            "Manicure",
                                            "Pedicure",
                                            "Facial",
                                            "Waxing",
                                            "Makeup Service"
                                           ],state="readonly",font=("Poppins",12),width=15)
service_list.grid(row=7,column=1,padx=10)
service_list.current(0)

service_label = tk.Label(genframe, text="Select Service:", font=("Poppins", 10, "italic"), bg="lightblue")
service_label.grid(row=6, column=1, columnspan=2)

# Date Entry    
date = tk.Label(genframe,text="Select Day:",font=("Poppins",10,"italic"),bg="lightblue")
date.grid(row=6,column=3,padx=10)
date_list = ttk.Combobox(genframe,values=["",
                                           "Monday",
                                           "Tuesday",
                                           "Wednesday",
                                           "Thursday",
                                           "Friday",
                                           "Saturday",
                                           ],state="readonly",font=("Poppins",12),width=15)
date_list.grid(row=7,column=3,padx=10)
date_list.current(0)

# Time Entry
select_time = tk.Label(genframe,text="Select Time:",font=("Poppins",10,"italic"),bg="lightblue")
select_time.grid(row=6,column=5,padx=10)
time_list = ttk.Combobox(genframe,values=["",
                                           "07:00am - 09:am",
                                           "09:00am - 11:00am",
                                           "BREAK TIME",
                                           "12:00pm - 2:00pm",
                                           "02:00pm - 04:00pm",
                                           "04:00pm - 6:00pm",
                                           "06:00pm - 08:00pm"],state="readonly",font=("Poppins",12),width=15)
time_list.grid(row=7,column=5,padx=10)
time_list.current(0)


# Buttons
submit_btn = tk.Button(window, text="Submit", font=("Poppins", 12, "bold"), bg="lightpink",command = appending)
submit_btn.grid(row=6, column=0, columnspan = 4, pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Poppins", 12, "bold"), bg="lightgreen",command = update)
update_btn.grid(row=6, column=1,columnspan = 4)

delete_btn = tk.Button(window, text="Delete", bg="red", fg="white",font=("Poppins", 12, "bold"),command = delete)
delete_btn.grid(row=6, column=2,columnspan = 5)

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Contact", "Service", "Day", "Time"),
    show="headings"
)

for headings in ("Order ID", "Customer Name", "Contact", "Service", "Day", "Time"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>",select_record)
display()
window.mainloop()